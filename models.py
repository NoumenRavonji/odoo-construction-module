
# -*- coding: utf-8 -*-
import openpyxl
from tempfile import TemporaryFile
from openerp import models, fields, api
from openerp.api import Environment as env
import openerp.addons.decimal_precision as dp
from openerp.tools.translate import _
from xlrd import open_workbook
import base64
from tempfile import TemporaryFile
import openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
import unicodedata
from openerp import http
from openerp.http import request
from openerp.addons.web.controllers.main import serialize_exception,content_disposition
from openerp.addons.sale.sale import sale_order 
from openerp import fields


class AvantMetre(models.Model):
	_inherit = 'mrp.bom'
	_name = "gent.avantmetre"

	# bom_line_ids =  fields.One2many('gent.avantmetre.line', 'bom_id', 'BoM Lines', copy=True)
	rubrique_line_ids =  fields.One2many('gent.avantmetre.rubrique', 'avantmetre_id', 'Rubrique', copy=True)
	nom = fields.Char(string="Nom")
	prix_total = fields.Float(string="Prix Total")
	
	state = fields.Selection([
        ('simple', "Simple"),
        ('avant_metre', "Avant Métré"),
        ('sous_detail', "Sous détail"),
    ], default='simple')

	partner_id=fields.Many2one('res.partner', 'Client', required=True, select=True)
	excel_avantmetre = fields.Binary(string='Excel File', store=True)

	@api.multi
	def import_excel_avant_metre(self):
		print "IMPORT EXCEL AVANT METRE"
		my_file = self.excel_avantmetre.decode('base64')
		excel_fileobj = TemporaryFile('wb+')
		excel_fileobj.write(my_file)
		excel_fileobj.seek(0)
		# Create workbook
		wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
		# Get the first sheet of excel file
		ws = wb[wb.get_sheet_names()[0]]
		start = False
		sections =[]
		self.rubrique_line_ids = []
		rubrique =""
		lines = []
		start_line =False
		for row in ws:
			val1 = row[0].value
			val2 = row[1].value
			val3 = row[2].value
			val4 = row[3].value
			val5 = row[4].value
			val6 = row[5].value
			if(val1 =="REF"):
				start = True
				continue
			if(start):
				print "start"
				print val1

				# SECTION

				if((val1 !="") and (val1 !=None) and (val2 ==None) and (val3 ==None) and (val4 ==None) and (val5 ==None) and (val6 ==None)):

					if(len(lines) > 0):
						print "here"
						print val1
						sections.append((0,0,{'rubrique': rubrique, "rubrique_bom_line_ids": lines}))
						lines =[]
						start_line = False
					start_line = True

					# if(self.env['gent.avantmetre.rubrique'].search([['rubrique','=',val1]])):
					# 	rubrique = self.env['gent.avantmetre.rubrique'].search([['rubrique','=',val1]]).rubrique
					# else:
					rubrique = val1

					

				# ADDING SECTION LINES
				if((val1 !=None) and (val2 !=None) and (val3 !=None) and (val4 !=None) and (val5 !=None) and (val6 !=None)):
					# ouvrage elementaire
					if(self.env['product.template'].search([['name',"=",val2],['gent_type', '=', 'ouvrage_elementaire']])):
						print "product exist"
					else:
						print "create product"
						print self.env['product.template'].search([['name',"=",val2],['gent_type', '=', 'ouvrage_elementaire']])
						self.env['product.template'].create({'name': val2, 'gent_type': 'ouvrage_elementaire'})

					product_id = self.env['product.product'].search([['name',"=",val2],["gent_type", "=","ouvrage_elementaire"]]).id

					# unite
					if(self.env['product.uom'].search([['name', "=",val3]])):
						print "uom exist"
					else:
						self.env['product.uom'].create({'name': val3, 'category_id': 1})

					product_uom = self.env['product.uom'].search([['name', "=",val3]])

					# qté
					# print "VALEURS"
					# print val2
					# print val3
					# print val4
					product_qty = float(val4)

					if(product_qty > 0):
						print "PRODUCT QUANT"
						print val1
						line = (0,0,{"product_id": product_id, "product_uom": product_uom, "product_qty": product_qty})
						print "LINE OK"
						print line
						lines.append(line)
					else:
						print product_qty


		self.rubrique_line_ids = sections
		pass


		# # ADDING SECTION LINES
		# 		if((val1 !=None) and (val2 !=None) and (val3 !=None) and (val4 !=None) and (val5 !=None) and (val6 !=None)):
		# 			# ouvrage elementaire
		# 			if(self.env['product.template'].search([['name',"=",val2],['gent_type', '=', 'ouvrage_elementaire']])):
		# 				print "product exist"
		# 			else:
		# 				print "create product"
		# 				self.env['product.product'].create({''})
		# 				print self.env['product.template'].search([['name',"=",val2],['gent_type', '=', 'ouvrage_elementaire']])
		# 				product_id = self.env['product.product'].search([['name',"=",val2]]).id
		# 				print "PRODUCT ID"
		# 				print product_id
		# 				print val2
		# 				self.env['product.template'].create({'name': val2, 'gent_type': 'ouvrage_elementaire'})

					
		# 			# print "PRODUCT ID"
		# 			# print product_id
		# 			# print val2
		# 			# unite
		# 			if(self.env['product.uom'].search([['name', "=",val3]])):
		# 				print "uom exist"
		# 			else:
		# 				self.env['product.uom'].create({'name': val3, 'category_id': 1})

		# 			product_uom = self.env['product.uom'].search([['name', "=",val3]])

		# 			# qté
		# 			print "VALEURS"
		# 			print val2
		# 			print val3
		# 			print val4
		# 			product_qty = float(val4)
		# 			if(product_qty > 0):
		# 				line = (0,0,{"product_id": product_id, "product_uom": product_uom, "product_qty": product_qty})
		# 				print "LINE OK"
		# 				print line
		# 				lines.append(line)

		# print "SECTIONS"
		# print sections
		# self.rubrique_line_ids = sections
		# pass

	@api.multi
	def strip_accents(self, text):
		return ''.join(c for c in unicodedata.normalize('NFKD', text) if unicodedata.category(c) != 'Mn')


	
	#rubrique_last = ""


	
	@api.model
	def create(self,vals,context=None):
		print "Produit"
		print vals
		product_id = vals['product_tmpl_id']
		print "hello"
		# ma_liste = vals['rubrique_line_ids'][0][2]['rubrique_bom_line_ids']
		# print vals['rubrique_line_ids'][0][2]['rubrique_bom_line_ids'][0][2]['product_id']
		try:
			ouv_elt = vals['rubrique_line_ids'][0][2]['rubrique_bom_line_ids'][0][2]['product_id']
			self.env['product.template'].browse([product_id]).write({'gent_type': 'chantier'})
			# self.env['product.template'].browse([ouv_elt]).write({'gent_type': 'ouvrage_elementaire'})
			for i in vals['rubrique_line_ids'][0][2]['rubrique_bom_line_ids']:
				j=i[2]['product_id']
				self.env['product.template'].browse([j]).write({'gent_type': 'ouvrage_elementaire'})
		except:
			print "IndexError"
		return super(AvantMetre,self).create(vals)



	@api.onchange("bom_line_ids")
	def bom_lines_change(self):
		print "CHANGE"
		rubrique_str = ""
		result =[]
		for line in self.bom_line_ids:
			rubrique_str = line.rubrique
			result.append((0,0,{'product_efficiency': line.product_efficiency, 'product_qty': line.product_qty, 'product_id': line.product_id, 'product_uom': line.product_uom, 'rubrique': line.rubrique}))
		# result.append((0,0,{'rubrique': rubrique_str, 'product_efficiency': 1.0, 'product_qty': 1.0, 'product_uom': 1}))
		#self.rubrique_last = rubrique_str
		#self.bom_line_ids = result

		pass

		# return super(AvantMetre, self).create(value)

class AvantMetreRubrique(models.Model):
	_name = "gent.avantmetre.rubrique"

	rubrique = fields.Char("Rubrique")
	rubrique_bom_line_ids = fields.One2many('gent.avantmetre.line', 'bom_id', 'BoM Lines', copy=True)
	avantmetre_id =  fields.Many2one('gent.avantmetre', 'Parent BoM', ondelete='cascade', select=True, required=True)

class AvantMetreLine(models.Model):
	_inherit = "mrp.bom.line"
	_name = 'gent.avantmetre.line'
	
	bom_id =  fields.Many2one('gent.avantmetre.rubrique', 'Parent BoM', ondelete='cascade', select=True, required=True)
	rubrique = fields.Char(string="Rubrique")

class Product2(models.Model):
	_inherit = "product.template"

	gent_type = fields.Selection([('chantier','Chantier'),('ouvrage_elementaire','Ouvrage Elementaire'),('composant_materiaux','Composant Materiaux'),('composant_materiel','Composant Materiel'),('composant_main_d_oeuvre','Composant main d\'oeuvre'),('epi','EPI'),('autres_charges','Autres charges')])
	gent_category = fields.Many2one(comodel_name='product.category', required=True,default=1)
	not_gent_product = fields.Boolean(default=1)
	
	@api.onchange('gent_category')
	def on_change_gent_category(self):
		print "Gent type CHANGE"
		self.categ_id = self.gent_category.id
		print self.gent_category.id
		# categ = self.browse(cr,uid,ids[0]).
		# categ_val=dict(self._columns['product_template'].selection).get(gent_type)
		for rec in self:
			# self.categ_id = rec.categ_id
			print rec.id

class Bde(models.Model):
	_inherit = "sale.order"
	
	avantmetre = fields.Many2one(comodel_name='gent.avantmetre', required=False)
	coeff= fields.Many2one(comodel_name="gent.coeff", string="Coefficient de vente K", store=True)
	is_bde = fields.Boolean("BDE", default=True)
	excel_devis = fields.Binary(string='Devis en Excel', store=True)
	
	state= fields.Selection([
			('bde', 'BDE'),
            ('draft', 'Draft Quotation'),
            ('sent', 'Quotation Sent'),
            ('cancel', 'Cancelled'),
            ('waiting_date', 'Waiting Schedule'),
            ('progress', 'Sales Order'),
            ('manual', 'Sale to Invoice'),
            ('shipping_except', 'Shipping Exception'),
            ('invoice_except', 'Invoice Exception'),
            ('done', 'Done'),
            ], string='Status', readonly=True, copy=False, help="Gives the status of the quotation or sales order.\
              \nThe exception status is automatically set when a cancel operation occurs \
              in the invoice validation (Invoice Exception) or in the picking list process (Shipping Exception).\nThe 'Waiting Schedule' status is set when the invoice is confirmed\
               but waiting for the scheduler to run on the order date.", select=True, default="bde")
	

	@api.multi
	def create_project(self):
		print "create"
		if(self.project == False):
			project_id = self.env['project.project'].create({'name': self.avantmetre.name, "partner_id": self.avantmetre.partner_id.id, "state": "open"  })
			self.project = True
			for order_line in self.order_line:
				self.env['project.task'].create({'name': order_line.name, 'project_id': project_id.id})


	@api.multi
	def import_excel_devis(self):
		print "IMPORT EXCEL DEVIS"
		my_file = self.excel_devis.decode('base64')
		excel_fileobj = TemporaryFile('wb+')
		excel_fileobj.write(my_file)
		excel_fileobj.seek(0)
		# Create workbook
		wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
		# Get the first sheet of excel file
		ws = wb[wb.get_sheet_names()[0]]
		start = False
		sections =[]
		aide = []
		self.order_line = []
		rubrique =""
		lines = []
		start_line =False
		for row in ws:
			val1 = row[0].value
			val2 = row[1].value
			val3 = row[2].value
			val4 = row[3].value
			val5 = row[4].value
			val6 = row[5].value
			if(val1 =="REF"):
				start = True
				continue
			if(start):
				print "start"

				# SECTION

				if((val1 !="") and (val1 !=None) and (val2 ==None) and (val3 ==None) and (val4 ==None) and (val5 ==None) and (val6 ==None)):
					print ".HERE WE ARE"
					if(self.env['sale_layout.category'].search([['name','=',val1]])):
						# rubrique = self.env['sale_layout.category'].search([['name','=',val1]]).id
						print "CATEGORY EXISTS"
						# print rubrique
					else:
						self.env['sale_layout.category'].create({'name': val1})
					
					rubrique = self.env['sale_layout.category'].search([['name','=',val1]]).id
					
					if(len(lines) > 0):
						sections.append((0,0,{'rubrique': rubrique, "rubrique_bom_line_ids": lines}))
						lines =[]
						start_line = False
					start_line = True

					# print rubrique	

				# ADDING SECTION LINES
				if((val1 !=None) and (val2 !=None) and (val3 !=None) and (val4 !=None) and (val5 !=None) and (val6 !=None)):
					# lines
					
					if(self.env['product.template'].search([['name',"=",val2],['gent_type', '=', 'ouvrage_elementaire']])):
						print "product exist"
					else:
						print "create product"
						if(self.env['product.template'].search([['name','=', val2],['gent_type','=','ouvrage_elementaire']])):
							print "product exists in product"
						else:
							self.env['product.template'].create({'name': val2, 'gent_type': 'ouvrage_elementaire', 'list_price': val5, 'not_gent_product':0})

					product_id = self.env['product.product'].search([['name',"=",val2],["gent_type", "=","ouvrage_elementaire"]]).id
					# print product_id
					# unite
					if(self.env['product.uom'].search([['name', "=",val3]])):
						print "uom exist"
					else:
						self.env['product.uom'].create({'name': val3, 'category_id': 1})

					product_uom = self.env['product.uom'].search([['name', "=",val3]])

					# qté
		
					product_qty = float(val4)

					# print "HERE"
					# print rubrique
					# print val4
					# print product_qty
					if(product_qty > 0):
						line = (0,0,{"product_id": product_id, "product_uom": product_uom, "product_uom_qty": product_qty, "prix_unit": float(val5), "sale_layout_cat_id": rubrique})
						print "LINE OK"
						print rubrique
						# print "prix_debourse"
						# print val5
						# print line
						print "150 VE?"
						print rubrique
						# print line
						lines.append(line)
						aide.append(line)
						# print lines
					print "INTERMEDIAIRE"	
					# print lines
				# aide.append(lines)
					# print aide
		# self.rubrique_line_ids = sections
		self.order_line = aide
		print "tapitra"
		print aide
		pass

	def onchange_pricelist_id(self, cr, uid, ids, pricelist_id, order_lines, context=None):
		value = {
            'currency_id': self.pool.get('product.pricelist').browse(cr, uid, pricelist_id, context=context).currency_id.id
		}
		return { 'value': value}

	@api.multi
	def convert_to_devis(self):
		self.state= "draft"

	@api.one
	@api.onchange('order_line')
	def on_change_order_line(self):
		print "CHANGE ORDER LINE"
		order_line_result = []
		for record in self:
			print "record"
			print record
			for line in record.order_line:
				somme_mo=0
				somme_materiel=0
				somme_materiaux=0

				for line2 in line.mo_line:
					somme_mo += line2.price_subtotal
				for line2 in line.materiel_line:
					somme_materiel += line2.price_subtotal
				for line2 in line.materiaux_line:
					print "MATERIAUX LINE MODIF"
					print line2.product_id
					somme_materiaux += line2.price_subtotal

			record.currency_id = self.env.ref('base.main_company').currency_id

	 # Form filling
	def unlink(self, cr, uid, ids, context=None):
		sale_orders = self.read(cr, uid, ids, ['state'], context=context)
		unlink_ids = []
		for s in sale_orders:
			if s['state'] in ['bde', 'draft', 'cancel']:
				unlink_ids.append(s['id'])
			else:
				raise osv.except_osv(_('Invalid Action!'), _('In order to delete a confirmed sales order, you must cancel it before!'))

		return super(sale_order, self).unlink(cr, uid, unlink_ids, context=context)

	def create(self, cr, uid, vals, context=None):
		print "BDE a Voir"
		print vals
		if context is None:
			context = {}
		if vals.get('partner_id'):
			print vals.get('partner_id')
		else:
			print self.pool.get('gent.avantmetre').browse(cr, uid,[vals['avantmetre']])
			vals['partner_id'] = self.pool.get('gent.avantmetre').browse(cr, uid,[vals['avantmetre']]).partner_id.id
		print "partner id"
		print vals.get('partner_id')
		if vals.get('name', '/') == '/':
			vals['name'] = self.pool.get('ir.sequence').get(cr, uid, 'sale.order', context=context) or '/'
		if vals.get('partner_id') and any(f not in vals for f in ['partner_invoice_id', 'partner_shipping_id', 'pricelist_id', 'fiscal_position']):
			defaults = self.onchange_partner_id(cr, uid, [], vals['partner_id'], context=context)['value']
			if not vals.get('fiscal_position') and vals.get('partner_shipping_id'):
				delivery_onchange = self.onchange_delivery_id(cr, uid, [], vals.get('company_id'), None, vals['partner_id'], vals.get('partner_shipping_id'), context=context)
				defaults.update(delivery_onchange['value'])
			vals = dict(defaults, **vals)
		ctx = dict(context or {}, mail_create_nolog=True)


		try:
			for k in range(0,len(vals['order_line'])):
				# print (vals['order_line'][k])
				try:
					for i in range (0,len(vals['order_line'][k][2]['materiel_line'])):
						if(len(vals['order_line'][k][2]['materiel_line'][i][2])!=0):
							print (vals['order_line'][k][2]['materiel_line'][i][2]['product_id'])
							mll=vals['order_line'][k][2]['materiel_line'][i][2]['product_id']
							self.pool.get('product.template').browse(cr,uid,mll).write({'gent_type': 'composant_materiel'})
				except:
					print ("no materiel line")
				try:
					for i in range (0,len(vals['order_line'][k][2]['materiaux_line'])):
						if(len(vals['order_line'][k][2]['materiaux_line'][i][2])!=0):
							print (vals['order_line'][k][2]['materiaux_line'][i][2]['product_id'])
							ml=vals['order_line'][k][2]['materiaux_line'][i][2]['product_id']
							self.pool.get('product.template').browse(cr,uid,ml).write({'gent_type': 'composant_materiaux'})
				except:
					print ("no materiaux line")
				try:
					for i in range (0,len(vals['order_line'][0][2]['mo_line'])):
						if(len(vals['order_line'][k][2]['mo_line'][i][2])!=0):
							print (vals['order_line'][k][2]['mo_line'][i][2]['product_id'])
							mo=vals['order_line'][k][2]['mo_line'][i][2]['product_id']
							self.pool.get('product.template').browse(cr,uid,mo).write({'gent_type': 'composant_main_d_oeuvre'})
							# self.pool.get('product.template').browse(cr,uid,mo).write({'categ_id':10})
							# self.pool.get('product.template').browse(cr,uid,mo).write({'gent_category':10})
				except:
					print ("no mo line")
		except:
			print ("no")
		new_id = super(Bde, self).create(cr, uid, vals, context=ctx)

		self.message_post(cr, uid, [new_id], body=_("Quotation created"), context=ctx)

		return new_id



	def write(self,cr, uid, ids,vals,context=None):
		print "Modifier"
		# self.button_dummy(cr,uid, ids,vals)
		print vals
		if('coeff' in vals):
			coeff = self.pool.get('gent.coeff').browse(cr,uid,[vals["coeff"]]).coeff
			order_line = self.pool.get("sale.order").browse(cr,uid,ids)
			print order_line
			for line in self.pool.get("sale.order").browse(cr,uid,ids).order_line:
				print "Line"
				print line
				line.write({
					'price_unit': line.prix_debourse * coeff
				})
				print line.price_unit

		if('order_line' in vals):
			for i in range(0,len(vals['order_line'])):
				if vals['order_line'][i][2] != False:
					try:
						for k in range(0,len(vals['order_line'][i][2]['materiaux_line'])):
							print vals['order_line'][i][2]['materiaux_line'][k][2]
							if vals['order_line'][i][2]['materiaux_line'][k][2] != False:
								mll = vals['order_line'][i][2]['materiaux_line'][k][2]['product_id']
								self.pool.get('product.template').browse(cr,uid,mll).write({'gent_type': 'composant_materiaux'})

					except:
						print("aucune modification materiaux_line")
					try:
						for k in range(0,len(vals['order_line'][i][2]['materiel_line'])):
							print vals['order_line'][i][2]['materiel_line'][k][2]
							if vals['order_line'][i][2]['materiel_line'][k][2] != False:
								ml= vals['order_line'][i][2]['materiel_line'][k][2]['product_id']
								self.pool.get('product.template').browse(cr,uid,ml).write({'gent_type': 'composant_materiel'})
					except:
						print("aucune modification materiel_line")
					try:
						for k in range(0,len(vals['order_line'][i][2]['mo_line'])):
							print vals['order_line'][i][2]['mo_line'][k][2]
							if vals['order_line'][i][2]['mo_line'][k][2] != False:
								mo = vals['order_line'][i][2]['mo_line'][k][2]['product_id'] 
								self.pool.get('product.template').browse(cr,uid,mo).write({'gent_type': 'composant_main_d_oeuvre'})
					except:
						print "aucune modification mo_line"
		return super(Bde, self).write(cr, uid,ids, vals, context)


	@api.onchange('avantmetre')
	def on_change_avantmetre(self):
		print "AVANT METRE CHANGE"
		result = []
		domain = []
		self.order_line = result
		print self.avantmetre.partner_id
		print "price_list"
		# print self.pricelist_id
		self.partner_id = self.avantmetre.partner_id
		for rubrique_line in self.avantmetre.rubrique_line_ids:
			rubrique = self.env['sale_layout.category'].create({"name": rubrique_line.rubrique, "sequence": 10})
			for line in rubrique_line.rubrique_bom_line_ids:
				# print line.product_id
				
				vals = self.pool.get('sale.order.line').product_id_change(self.env.cr, self.env.uid, [], self.pricelist_id.id, line.product_id.id, line.product_qty, line.product_uom.id, 0, False, '', self.partner_id.id)
				vals['value'].update({
					'delay': 0,
					'state': "draft",
	              'product_id': line.product_id.id,
	              'product_uom': line.product_uom,
	              'product_uom_qty': line.product_qty,
	              'price_unit': 0,
	              'sale_layout_cat_id': rubrique
	            })
				result.append(vals['value'])
				
		self.order_line = result

	@api.one
	def button_dummy(self, vals):
		res = super(Bde, self).button_dummy()
		print "button_dummy vals"
		print vals
		if(self.coeff.coeff):
			for record in self:
				for line in record.order_line:
					print "BUTTON DUMMY"
					line.price_unit = line.prix_debourse * self.coeff.coeff
					print line.price_unit
		return res


class GentSaleOrderLine(models.Model):
	_inherit = "sale.order.line"

	ouvrage_elementaire = fields.Many2many('gent.ouvrage.elementaire', string='Ouvrage élémentaire', store=True)

	prix_debourse = fields.Float('Prix déboursé', store=True, compute="_compute_oe_pu")

	mo_line = fields.One2many('gent.bde.composant', 'gent_mo_order_line_id', "Main d'oeuvre")
	autres_charges_line = fields.One2many('gent.bde.composant', 'gent_mo_order_line_id', "Main d'oeuvre")
	epi_line = fields.One2many('gent.bde.composant', 'gent_epi_order_line_id', "EPI")
	materiel_line = fields.One2many('gent.bde.composant', 'gent_materiel_order_line_id', "Matériels")
	materiaux_line = fields.One2many('gent.bde.composant', 'gent_materiaux_order_line_id', "Matériaux")

	# rendement = fields.Float('Rendement',default=1,required=True)
	mo_lines_subtotal = fields.Float('Total')
	price_unit= fields.Float('Unit Price', digits_compute= dp.get_precision('Product Price'), store=True, readonly=True)

	price_subtotal = fields.Float('Montant', digits_compute= dp.get_precision('Product Price'), store=True, readonly=True,compute='_compute_order_line_montant')

	product_id =  fields.Many2one('product.product', 'Product', domain=[('sale_ok', '=', True), ('gent_type', '=', 'ouvrage_elementaire')], change_default=True, readonly=True, states={'draft': [('readonly', False)]}, ondelete='restrict')
	product_uom_qty =fields.Float('Quantity', digits_compute= dp.get_precision('Product UoS'), required=True, readonly=True, states={'draft': [('readonly', False)],'bde': [('readonly', False)]})
        

	@api.depends('price_unit','product_uom_qty')
	def _compute_order_line_montant(self):
		for record in self:
			record.price_subtotal= record.price_unit * record.product_uom_qty

	# @api.depends('ouvrage_elementaire')
	# def _compute_rendement(self):
	# 	for record in self:
	# 		record.rendement = record.ouvrage_elementaire.rendement
	@api.one
	@api.depends('mo_line','materiaux_line','materiel_line')
	def _compute_oe_pu(self):
		somme_mo=0
		somme_materiel=0
		somme_materiaux=0
		# try:
		# 	for record in self:
		# 		somme_mo = sum(record.price_subtotal for line in record.mo_line)
		# 		somme_materiaux = sum(record.price_subtotal for line in record.materiel_line)
		# 		somme_materiel = sum(record.price_subtotal for line in record.materiaux_line)
		# except:
		# 	print "Erreur d'index"
		for record in self:
			print record
			for line in record.mo_line:
				somme_mo += line.price_subtotal
			for line in record.materiel_line:
				somme_materiel += line.price_subtotal
			for line in record.materiaux_line:
				somme_materiaux += line.price_subtotal
		self.prix_debourse = (somme_mo + somme_materiel + somme_materiaux)*self.product_uom_qty
		self.price_unit = (somme_mo + somme_materiel + somme_materiaux)*self.product_uom_qty
		# if self.ouvrage_elementaire.rendement == 0:
		# 	self.prix_debourse = (somme_mo + somme_materiel + somme_materiaux)
		# 	self.price_unit = (somme_mo + somme_materiel + somme_materiaux)
		# else:
		# 	# self.prix_debourse = (somme_mo + somme_materiel + somme_materiaux)
		# 	self.prix_debourse = (somme_mo + somme_materiel + somme_materiaux)/self.ouvrage_elementaire.rendement
		# 	self.price_unit = (somme_mo + somme_materiel + somme_materiaux)/self.ouvrage_elementaire.rendement

		print "BDE PRIX DEBOURSE"
		print self.prix_debourse
		print self.price_unit
		# id_ouvrage = self.ouvrage_elementaire.rendement
		# if id_ouvrage != 0:
		# 	print self.price_unit/id_ouvrage

	
	def create(self, cr, uid, values, context=None):
		print "CREATING SALE ORDER"
		if values.get('order_id') and values.get('product_id') and  any(f not in values for f in ['name', 'price_unit', 'product_uom_qty', 'product_uom']):
			order = self.pool['sale.order'].read(cr, uid, values['order_id'], ['pricelist_id', 'partner_id', 'date_order', 'fiscal_position'], context=context)
			defaults = self.product_id_change(cr, uid, [], order['pricelist_id'][0], values['product_id'],
				qty=float(values.get('product_uom_qty', False)),
				uom=values.get('product_uom', False),
				qty_uos=float(values.get('product_uos_qty', False)),
				uos=values.get('product_uos', False),
				name=values.get('name', False),
				partner_id=order['partner_id'][0],
				date_order=order['date_order'],
				fiscal_position=order['fiscal_position'][0] if order['fiscal_position'] else False,
				flag=False,  # Force name update
                context=dict(context or {}, company_id=values.get('company_id'))
			)['value']
			if defaults.get('tax_id'):
				defaults['tax_id'] = [[6, 0, defaults['tax_id']]]
			values = dict(defaults, **values)
		return super(GentSaleOrderLine, self).create(cr, uid, values, context=context)

	@api.onchange('ouvrage_elementaire')
	def on_change_ouvrage_elementaire(self):
		print "CHANGE"
		self.mo_line = []
		self.materiaux_line = []
		self.materiel_line = []
		if(len(self.ouvrage_elementaire) > 0):
			
			for ouvrage in self.ouvrage_elementaire:
				for line in ouvrage.mo_line:
					self.mo_line +=self.mo_line.new({
								'product_id': line.product_id,
								'price_unit': line.price_unit,
								'price_subtotal': line.price_subtotal,
								'product_uom_qty': line.product_uom_qty,
								'product_uom': line.product_uom,
								'gent_mo_order_line_id': self.id

							})
			for ouvrage in self.ouvrage_elementaire:
				for line in ouvrage.materiaux_line:
					self.materiaux_line +=self.materiaux_line.new({
								'product_id': line.product_id,
								'price_unit': line.price_unit,
								'price_subtotal': line.price_subtotal,
								'product_uom_qty': line.product_uom_qty,
								'product_uom': line.product_uom,
								'gent_materiaux_order_line_id': self.id

						})
			for ouvrage in self.ouvrage_elementaire:
				for line in ouvrage.materiel_line:
					self.materiel_line +=self.materiel_line.new({
								'product_id': line.product_id,
								'price_unit': line.price_unit,
								'price_subtotal': line.price_subtotal,
								'product_uom_qty': line.product_uom_qty,
								'product_uom': line.product_uom,
								'gent_materiel_order_line_id': self.id

							})


class BdeLine(models.Model):
	_name = 'gent.bde.composant'


	gent_mo_order_line_id = fields.Many2one('sale.order.line', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_autres_charges_order_line_id = fields.Many2one('sale.order.line', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_materiel_order_line_id = fields.Many2one('sale.order.line', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_materiaux_order_line_id = fields.Many2one('sale.order.line', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_epi_order_line_id = fields.Many2one('sale.order.line','Parent Order Line' , ondelete='cascade',select=True, readonly=True)

	gent_oe_mo_line_id = fields.Many2one('gent.ouvrage.elementaire', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_oe_materiel_id = fields.Many2one('gent.ouvrage.elementaire', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_oe_materaux_id = fields.Many2one('gent.ouvrage.elementaire', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_oe_epi_line_id = fields.Many2one('gent.ouvrage.elementaire', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)
	gent_oe_autres_charges_line_id = fields.Many2one('gent.ouvrage.elementaire', 'Parent Order Line',ondelete='cascade', select=True, readonly=True)


	product_id =  fields.Many2one('product.product', 'Product', domain=[('sale_ok', '=', True)], change_default=True, ondelete='cascade', required=True)
	price_unit = fields.Float('Unit Price', required=True, digits_compute= dp.get_precision('Product Price'))



	price_subtotal =  fields.Float('Montant',compute='_compute_subtotal')
	product_uom_qty =  fields.Float('Quantity', digits_compute= dp.get_precision('Product UoS'), required=True)
	product_uom = fields.Many2one('product.uom', 'Unit of Measure ', required=True)

	@api.one
	@api.depends('price_unit','product_uom_qty')
	def _compute_subtotal(self):
		for record in self:
			record.price_subtotal = record.price_unit * record.product_uom_qty


class OuvrageElementaire(models.Model):
	_name = "gent.ouvrage.elementaire.import"
	excel_file = fields.Binary(string='Excel File')
	# example_file = fields.Binary(string='File')

	# @api.multi
	def _get_default_image(self,cr,uid,context=None):
		attach_obj = self.pool.get('ir.attachment')
		attach_data_ids=attach_obj.search(cr,uid,[('name','=','Gent_ouvrage_modele')])
		attach_data = attach_obj.browse(cr,uid,attach_data_ids)
		print "Ito eh"
		print attach_data_ids
		return attach_data.datas
	_defaults={'excel_file': _get_default_image}
	# _columns = {
	# 	'excel_file' : fields.binary('Welcome Letter'),
	# }
	# _defaults={
	# 	'excel_file' : _get_default_welcome_letter 
	# }

	
	@api.multi
	def import_excel(self):
		print "IMPORT EXCEL"
		# Generating of the excel file to be read by openpyxl
		my_file = self.excel_file.decode('base64')
		excel_fileobj = TemporaryFile('wb+')
		excel_fileobj.write(my_file)
		excel_fileobj.seek(0)

		# Create workbook
		wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
		# Get the first sheet of excel file
		ws = wb[wb.get_sheet_names()[0]]

		print "DISPLAYING EXCEL"
		a=list()
		d=dict()
		c=0
		# if (ws.cell(row=1,column=1).value == None):
		for row in ws:
			if(row[0].value=="DESIGNATION" and row[2].value==None):
				d = {}
			if(row[0].value=="DESIGNATION" and row[2].value!=None):
				d={}
				key=str()
				d['nom_section']=row[2].value
				d['MATERIAUX']=list()
				d['MATERIELS']=list()
				d['MO']=list()
			if(row[0].value=="A -MAIN D'OEUVRE"):
				key='MO'
				continue
			if(row[0].value=="B - MATERIAUX"):
				key='MATERIAUX'
				continue
			if(row[0].value=="C - MATERIEL - OUTILLAGE"):
				key='MATERIELS'
				continue
			if(row[1].value==None):
				continue
			if(row[3].value==0):
				continue
			if(row[1].value == "DESIGNATION" and row[2].value=="U" and row[3].value=="QUANTITES" and row[4].value=="PRIX UNITAIRE"):
				continue
			try:
				d[key].append({'Designation':row[1].value,'Unite':row[2].value,'Quantite':row[3].value,'Prix Unitaire':row[4].value})
			except:
				pass
			
			# Insertion de la section dans la liste
			if(d in a):
				continue
			if(d=={}):
				continue
			else:
				a.append(d)
		# if(ws.cell(row=1,column=1).value == "N DE PRIX:001"):
		# 	print "Vos"
		# 	for row in ws:
		# 		if(row[0].value=="N DE PRIX:001"):
		# 				d={}
		# 				cle=str()
		# 				d['nom_section']=row[1].value
		# 				d['MATERIAUX']=list()
		# 				d['MATERIELS']=list()
		# 				d['MO']=list()
		# 				# d['rendement']=0
		# 		if row[0].value in ['TOTAL MATERIAUX','TOTAL MATERIELS','DESIGNATION']:
		# 			continue
		# 		if(row[0].value=="MATERIAUX"):
		# 			cle='MATERIAUX'
		# 			continue
		# 		if(row[0].value=="MATERIELS"):
		# 			cle='MATERIELS'
		# 			continue

				
		# 		if(row[0].value != None):
		# 			if((row[0].value == "MAIN D'OEUVRE") or (row[0].value.encode('ascii', 'xmlcharrefreplace') == "MAIN D'&#338;UVRE" )):
		# 				cle='MO'
		# 				continue
		# 		if(row[0].value==None):
		# 			if(row[3].value==None):
		# 				continue	
		# 			else:
		# 				row[0].value=cle+'-'+d['nom_section']
		# 		if(row[0].value=="TOTAL MAIN D\'OEUVRE"):
		# 			# cle=str()
		# 			continue
		# 		if(row[0].value=='Coeficient K='):
		# 			continue
		# 		# if(row[0].value=="Rendement R="):
		# 		# 	d['rendement']=row[1].value
		# 		# 	cle=""
		# 			continue
		# 		if(row[2].value==None and row[3].value==None):
		# 			continue
		# 		if(row[1].value==None):
		# 			row[1].value='u'
		# 		# if(row[1].value==None):
		# 		# 	continue
		# 		try:
		# 			d[cle].append({'Designation':row[0].value,'Unite':row[1].value,'Quantite':row[2].value,'Prix Unitaire':row[3].value})
		# 		except KeyError:
		# 			pass
				
		# 		# Insertion de la section dans la liste
		# 		if(d in a):
		# 			continue
		# 		else:
		# 			a.append(d)

		for ouvrage in a:
			if (self.env['product.product'].search([['name', '=',ouvrage['nom_section']]])):
				print "Bonjour! :",self.env['product.product'].search([('name','=',ouvrage['nom_section'])]).id
			else:
				self.env['product.template'].create({'name':ouvrage['nom_section'],'gent_type':'ouvrage_elementaire','not_gent_product':0})
			id_ouv = self.env['product.product'].search([('name','=',ouvrage['nom_section'])]).id
			if (self.env['gent.ouvrage.elementaire'].search([['product_id','=',id_ouv]])):
				print "ouvrage elementaire existant"
			else:
				self.env['gent.ouvrage.elementaire'].create({'product_id':id_ouv,'product_uom':1})
			id_line = self.env['gent.ouvrage.elementaire'].search([['product_id','=',id_ouv]])
			print ouvrage['nom_section'], ":",id_line.id

			for materiaux in ouvrage['MATERIAUX']:
				if(self.env['product.category'].search([('name','=',"Composant Materiaux")])):
					id_gent_category = self.env['product.category'].search([('name','=',"Composant Materiaux")]).id
					# print id_unit_mo
				else:
					# print "unite qui n'existe pas encore"
					self.env['product.category'].create({'name':"Composant Materiaux"})
					id_gent_category = self.env['product.category'].search([('name','=',"Composant Materiaux")]).id
				print "MATERIAUX ",materiaux['Designation']
				if materiaux['Unite'] == 'u':
					materiaux['Unite']='Unité(s)'
				elif materiaux['Unite'] == 'l':
					materiaux['Unite']='Litre(s)'
				elif materiaux['Unite'] == 'Fft':
					materiaux['Unite'] = 'fft'
				elif materiaux['Unite'] == None:
					continue
				if(self.env['product.uom'].search([['name','=',materiaux['Unite']]])):
					id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
					print id_unit_mo
				else:
					print "unite qui n'existe pas encore"
					self.env['product.uom'].create({'name':materiaux['Unite'],'category_id':1})
				if(self.env['product.product'].search([['name','=',materiaux['Designation']]])):
					print "Le produit existe deja"
				else:
					self.env['product.template'].create({'name':materiaux['Designation'],'gent_type':'composant_materiaux','categ_id':id_gent_category,'gent_category':id_gent_category, 'standard_price': materiaux['Prix Unitaire'], 'sale_ok': False,'type': 'product','not_gent_product':0 })
				id_mo = self.env['product.product'].search([('name','=',materiaux['Designation'])]).id
				print id_mo
				id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
				print "ID_LINE : ",id_line.id
				
				if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',materiaux['Prix Unitaire']],['product_uom_qty','=',materiaux['Quantite']],['gent_oe_materaux_id','=',id_line.id]])):
					print "BDE_composant existant"
				else:
					self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':materiaux['Prix Unitaire'],'product_uom_qty':materiaux['Quantite'],'gent_oe_materaux_id':id_line.id})
			for materiel in ouvrage['MATERIELS']:
				# id_gent_category = self.env['product.category'].search([('name','=',"Composant Materiel")]).id
				if(self.env['product.category'].search([('name','=',"Composant Materiel")])):
					id_gent_category = self.env['product.category'].search([('name','=',"Composant Materiel")]).id
					# print id_unit_mo
				else:
					# print "unite qui n'existe pas encore"
					self.env['product.category'].create({'name':"Composant Materiel"})
					id_gent_category = self.env['product.category'].search([('name','=',"Composant Materiel")]).id
				if materiel['Unite'] == 'u':
					materiel['Unite']='Unité(s)'
				elif materiel['Unite'] == 'l':
					materiel['Unite']='Litre(s)'
				elif materiel['Unite'] == 'Fft':
					materiel['Unite'] = 'fft'
				elif materiel['Unite'] == None:
					continue
				if(self.env['product.uom'].search([['name','=',materiel['Unite']]])):
					id_unit_mo = self.env['product.uom'].search([('name','=',materiel['Unite'])]).id
					print id_unit_mo
				else:
					print "unite qui n'existe pas encore"
					self.env['product.uom'].create({'name':materiel['Unite'],'category_id':1})
				if(self.env['product.product'].search([['name','=',materiel['Designation']]])):
					print "Le produit existe deja"
				else:
					self.env['product.template'].create({'name':materiel['Designation'],'gent_type':'composant_materiel','categ_id':id_gent_category,'gent_category':id_gent_category, 'standard_price': materiel['Prix Unitaire'], 'sale_ok': False,'type': 'product','not_gent_product':0})
				id_mo = self.env['product.product'].search([('name','=',materiel['Designation'])]).id
				print id_mo
				id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
				print "ID_LINE : ",id_line.id
				
				if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',materiel['Prix Unitaire']],['product_uom_qty','=',materiel['Quantite']],['gent_oe_materiel_id','=',id_line.id]])):
					print "BDE_composant existant"
				else:
					self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':materiel['Prix Unitaire'],'product_uom_qty':materiel['Quantite'],'gent_oe_materiel_id':id_line.id})
			for mo in ouvrage['MO']:
				if(self.env['product.category'].search([('name','=',"Composant main d'oeuvre")])):
					id_gent_category = self.env['product.category'].search([('name','=',"Composant main d'oeuvre")]).id
					# print id_unit_mo
				else:
					# print "unite qui n'existe pas encore"
					self.env['product.category'].create({'name':"Composant main d'oeuvre"})
					id_gent_category = self.env['product.category'].search([('name','=',"Composant main d'oeuvre")]).id
				# id_gent_category = self.env['product.category'].search([('name','=',"Composant main d'oeuvre")]).id
				print "MO ",mo['Designation']
				if mo['Unite'] in ['','u','u ']:
					mo['Unite']='Unité(s)'
				elif mo['Unite'] == 'l':
					mo['Unite']='Litre(s)'
				elif mo['Unite'] == 'Fft':
					mo['Unite'] = 'fft'
				elif mo['Unite'] == None:
					continue
				if(self.env['product.uom'].search([['name','=',mo['Unite']]])):
					id_unit_mo = self.env['product.uom'].search([('name','=',mo['Unite'])]).id
					print id_unit_mo
				else:
					print "unite qui n'existe pas encore"
					self.env['product.uom'].create({'name':mo['Unite'],'category_id':1})
				if(self.env['product.product'].search([['name','=',mo['Designation']]])):
					print "Le produit existe deja"
				else:
					self.env['product.template'].create({'name':mo['Designation'],'gent_type':'composant_main_d_oeuvre','categ_id':id_gent_category,'gent_category':id_gent_category, 'standard_price': mo['Prix Unitaire'], 'sale_ok': False,'type': 'service','not_gent_product':0})
				id_mo = self.env['product.product'].search([('name','=',mo['Designation'])]).id
				print id_mo
				id_unit_mo = self.env['product.uom'].search([('name','=',mo['Unite'])]).id
				print "ID_LINE : ",id_line.id
				print "HELLO"
				if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',mo['Prix Unitaire']],['product_uom_qty','=',mo['Quantite']],['gent_oe_mo_line_id','=',id_line.id]])):
					print "BDE_composant existant"
				else:
					self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':mo['Prix Unitaire'],'product_uom_qty':mo['Quantite'],'gent_oe_mo_line_id':id_line.id})
	#nouvel_import
		# print "DISPLAYING EXCEL"
		# 	a=list()
		# 	d=dict()
		# 	c=0
		# 	cle=str()
		# 	for row in ws:
		# 		print "ENCODING"
		# 		print row[0].value
		# 		if(row[0].value=="N DE PRIX:001"):
		# 			d={}
		# 			cle=str()
		# 			d['nom_section']=row[1].value
		# 			d['MATERIAUX']=list()
		# 			d['MATERIELS']=list()
		# 			d['MO']=list()
		# 			d['rendement']=0
		# 		if row[0].value in ['TOTAL MATERIAUX','TOTAL MATERIELS','DESIGNATION']:
		# 			continue
		# 		if(row[0].value=="MATERIAUX"):
		# 			cle='MATERIAUX'
		# 			continue
		# 		if(row[0].value=="MATERIELS"):
		# 			cle='MATERIELS'
		# 			continue

				
		# 		if(row[0].value != None):
		# 			if((row[0].value == "MAIN D'OEUVRE") or (row[0].value.encode('ascii', 'xmlcharrefreplace') == "MAIN D'&#338;UVRE" )):
		# 				cle='MO'
		# 				continue
		# 		if(row[0].value==None):
		# 			if(row[3].value==None):
		# 				continue	
		# 			else:
		# 				row[0].value=cle+'-'+d['nom_section']
		# 		if(row[0].value=="TOTAL MAIN D\'OEUVRE"):
		# 			# cle=str()
		# 			continue
		# 		if(row[0].value=='Coeficient K='):
		# 			continue
		# 		if(row[0].value=="Rendement R="):
		# 			d['rendement']=row[1].value
		# 			cle=""
		# 			continue
		# 		if(row[2].value==None and row[3].value==None):
		# 			continue
		# 		if(row[1].value==None):
		# 			row[1].value='u'
		# 		# if(row[1].value==None):
		# 		# 	continue
		# 		try:
		# 			d[cle].append({'Designation':row[0].value,'Unite':row[1].value,'Quantite':row[2].value,'Prix Unitaire':row[3].value})
		# 		except KeyError:
		# 			pass
				
		# 		# Insertion de la section dans la liste
		# 		if(d in a):
		# 			continue
		# 		else:
		# 			a.append(d)
		# 		# print d['nom_section']
		# 		print a[0]['rendement']
		# 	for ouvrage in a:
		# 		if (self.env['product.product'].search([['name', '=',ouvrage['nom_section']]])):
		# 			print "Bonjour! :",self.env['product.product'].search([('name','=',ouvrage['nom_section'])]).id
		# 		else:
		# 			self.env['product.template'].create({'name':ouvrage['nom_section'],'gent_type':'ouvrage_elementaire'})
		# 		id_ouv = self.env['product.product'].search([('name','=',ouvrage['nom_section'])]).id
		# 		if (self.env['gent.ouvrage.elementaire'].search([['product_id','=',id_ouv]])):
		# 			print "ouvrage elementaire existant"
		# 		else:
		# 			self.env['gent.ouvrage.elementaire'].create({'product_id':id_ouv,'product_uom':1,'rendement':ouvrage['rendement']})
		# 		id_line = self.env['gent.ouvrage.elementaire'].search([['product_id','=',id_ouv]])
		# 		print ouvrage['nom_section'], ":",id_line.id

		# 		for materiaux in ouvrage['MATERIAUX']:
		# 			print "MATERIAUX ",materiaux['Designation']
		# 			if materiaux['Unite'] == 'u':
		# 				materiaux['Unite']='Unité(s)'
		# 			elif materiaux['Unite'] == 'l':
		# 				materiaux['Unite']='Litre(s)'
		# 			elif materiaux['Unite'] == 'Fft':
		# 				materiaux['Unite'] = 'fft'
		# 			elif materiaux['Unite'] == None:
		# 				continue
		# 			if(self.env['product.uom'].search([['name','=',materiaux['Unite']]])):
		# 				id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
		# 				print id_unit_mo
		# 			else:
		# 				print "unite qui n'existe pas encore"
		# 				self.env['product.uom'].create({'name':materiaux['Unite'],'category_id':1})
		# 			if(self.env['product.product'].search([['name','=',materiaux['Designation']]])):
		# 				print "Le produit existe deja"
		# 			else:
		# 				self.env['product.template'].create({'name':materiaux['Designation'],'gent_type':'composant_materiaux'})
		# 			id_mo = self.env['product.product'].search([('name','=',materiaux['Designation'])]).id
		# 			print id_mo
		# 			id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
		# 			print "ID_LINE : ",id_line.id
					
		# 			if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',materiaux['Prix Unitaire']],['product_uom_qty','=',materiaux['Quantite']],['gent_oe_materaux_id','=',id_line.id]])):
		# 				print "BDE_composant existant"
		# 			else:
		# 				self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':materiaux['Prix Unitaire'],'product_uom_qty':materiaux['Quantite'],'gent_oe_materaux_id':id_line.id})
		# 		for materiel in ouvrage['MATERIELS']:

		# 			if materiel['Unite'] == 'u':
		# 				materiel['Unite']='Unité(s)'
		# 			elif materiel['Unite'] == 'l':
		# 				materiel['Unite']='Litre(s)'
		# 			elif materiel['Unite'] == 'Fft':
		# 				materiel['Unite'] = 'fft'
		# 			elif materiel['Unite'] == None:
		# 				continue
		# 			if(self.env['product.uom'].search([['name','=',materiel['Unite']]])):
		# 				id_unit_mo = self.env['product.uom'].search([('name','=',materiel['Unite'])]).id
		# 				print id_unit_mo
		# 			else:
		# 				print "unite qui n'existe pas encore"
		# 				self.env['product.uom'].create({'name':materiel['Unite'],'category_id':1})
		# 			if(self.env['product.product'].search([['name','=',materiel['Designation']]])):
		# 				print "Le produit existe deja"
		# 			else:
		# 				self.env['product.template'].create({'name':materiel['Designation'],'gent_type':'composant_materiel'})
		# 			id_mo = self.env['product.product'].search([('name','=',materiel['Designation'])]).id
		# 			print id_mo
		# 			id_unit_mo = self.env['product.uom'].search([('name','=',materiaux['Unite'])]).id
		# 			print "ID_LINE : ",id_line.id
					
		# 			if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',materiel['Prix Unitaire']],['product_uom_qty','=',materiel['Quantite']],['gent_oe_materiel_id','=',id_line.id]])):
		# 				print "BDE_composant existant"
		# 			else:
		# 				self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':materiel['Prix Unitaire'],'product_uom_qty':materiel['Quantite'],'gent_oe_materiel_id':id_line.id})
		# 		for mo in ouvrage['MO']:
		# 			print "MO ",mo['Designation']
		# 			if mo['Unite'] in ['','u','u ']:
		# 				mo['Unite']='Unité(s)'
		# 			elif mo['Unite'] == 'l':
		# 				mo['Unite']='Litre(s)'
		# 			elif mo['Unite'] == 'Fft':
		# 				mo['Unite'] = 'fft'
		# 			elif mo['Unite'] == None:
		# 				continue
		# 			if(self.env['product.uom'].search([['name','=',mo['Unite']]])):
		# 				id_unit_mo = self.env['product.uom'].search([('name','=',mo['Unite'])]).id
		# 				print id_unit_mo
		# 			else:
		# 				print "unite qui n'existe pas encore"
		# 				self.env['product.uom'].create({'name':mo['Unite'],'category_id':1})
		# 			if(self.env['product.product'].search([['name','=',mo['Designation']]])):
		# 				print "Le produit existe deja"
		# 			else:
		# 				self.env['product.template'].create({'name':mo['Designation'],'gent_type':'composant_main_d_oeuvre'})
		# 			id_mo = self.env['product.product'].search([('name','=',mo['Designation'])]).id
		# 			print id_mo
		# 			id_unit_mo = self.env['product.uom'].search([('name','=',mo['Unite'])]).id
		# 			print "ID_LINE : ",id_line.id
		# 			print "HELLO"
		# 			if(self.env['gent.bde.composant'].search([['product_id','=',id_mo],['product_uom','=',id_unit_mo],['price_unit','=',mo['Prix Unitaire']],['product_uom_qty','=',mo['Quantite']],['gent_oe_mo_line_id','=',id_line.id]])):
		# 				print "BDE_composant existant"
		# 			else:
		# 				self.env['gent.bde.composant'].create({'product_id':id_mo,'product_uom':id_unit_mo,'price_unit':mo['Prix Unitaire'],'product_uom_qty':mo['Quantite'],'gent_oe_mo_line_id':id_line.id})


	def strip_accents(self, text):
		return ''.join(c for c in unicodedata.normalize('NFKD', text) if unicodedata.category(c) != 'Mn')                    

	
class OuvrageElementaire(models.Model):
	_name = "gent.ouvrage.elementaire"
	_rec_name="product_id"

	product_id =  fields.Many2one('product.product', 'Product', domain=[('sale_ok', '=', True), ('gent_type', '=', 'ouvrage_elementaire')], change_default=True, required=True, ondelete='restrict')

	rendement = fields.Float('Rendement',default=1,required=True)
	prix_debourse = fields.Float('Prix déboursé', store=True, compute='_compute_oe_pu')

	mo_line = fields.One2many('gent.bde.composant', 'gent_oe_mo_line_id', "Main d'oeuvre", copy=True)
	autres_charges_line = fields.One2many('gent.bde.composant', 'gent_oe_autres_charges_line_id', "Autres charges", copy=True)
	epi_line = fields.One2many('gent.bde.composant', 'gent_oe_epi_line_id', "EPI", copy=True)
	materiel_line = fields.One2many('gent.bde.composant', 'gent_oe_materiel_id', "Matériels", copy=True)
	materiaux_line = fields.One2many('gent.bde.composant', 'gent_oe_materaux_id', "Matériaux", copy=True)
	price_unit= fields.Float('Unit Price', store=True, readonly=True)
	product_uom_qty= fields.Float('Quantity',default=1, digits_compute= dp.get_precision('Product UoS'), required=True, readonly=True)
	product_uom =  fields.Many2one('product.uom', 'Unit of Measure ', required=True)

	price_subtotal = fields.Float('Montant', digits_compute= dp.get_precision('Product Price'), store=True, readonly=True,compute='_compute_order_line_montant')
	

	# @api.multi
	# def import_excel(self):
	# 	print "IMPORT EXCEL"
	# 	# Generating of the excel file to be read by openpyxl
	# 	my_file = self.excel_file.decode('base64')
	# 	excel_fileobj = TemporaryFile('wb+')
	# 	excel_fileobj.write(my_file)
	# 	excel_fileobj.seek(0)

	# 	# Create workbook
	# 	wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
	# 	# Get the first sheet of excel file
	# 	ws = wb[wb.get_sheet_names()[0]]

	# 	print "DISPLAYING EXCEL"
	# 	# Iteration on each rows in excel
	# 	a=list()
	# 	d=dict()
	# 	c=0
	# 	cle=str()
	# 	for row in ws:
	# 		if(row[0].value=="N DE PRIX:001"):
	# 			d={}
	# 			cle=str()
	# 			d['nom_section']=row[1].value
	# 			d['MATERIAUX']=list()
	# 			d['MATERIELS']=list()
	# 			d['MO']=list()
	# 		if row[0].value in ['TOTAL MATERIAUX','TOTAL MATERIELS','TOTAL MAIN D\'OEUVRE','DESIGNATION']:
	# 			continue
	# 		if(row[0].value=="MATERIAUX"):
	# 			cle='MATERIAUX'
	# 			continue
	# 		if(row[0].value=="MATERIELS"):
	# 			cle='MATERIELS'
	# 			continue
	# 		if(row[0].value=="MAIN D'OEUVRE"):
	# 			cle='MO'
	# 			continue
	# 		if(row[0].value=="TOTAL MAIN D\'OEUVRE"):
	# 			cle=str()
	# 			continue
	# 		try:
	# 			d[cle].append({'Designation':row[0].value,'Unite':row[1].value,'Quantite':row[2].value,'Prix Unitaire':row[3].value})
	# 		except KeyError:
	# 			pass
			
	# 		# Insertion de la section dans la liste
	# 		if(d in a):
	# 			continue
	# 		else:
	# 			a.append(d)
	# 		print a
		# self.env['product.template'].create({'name':'produit_valeur'})
		

	excel_file = fields.Binary(string='Excel File')

	# @api.one
	@api.depends('mo_line','materiaux_line','materiel_line')
	def _compute_oe_pu(self):
		somme_mo=0
		somme_materiel=0
		somme_materiaux=0
		# try:
		# 	for record in self:
		# 		somme_mo = sum(record.price_subtotal for line in record.mo_line)
		# 		somme_materiaux = sum(record.price_subtotal for line in record.materiel_line)
		# 		somme_materiel = sum(record.price_subtotal for line in record.materiaux_line)
		# except:
		# 	print "Erreur d'index"
		for record in self:
			for line in record.mo_line:
				somme_mo += line.price_subtotal
			for line in record.materiel_line:
				somme_materiel += line.price_subtotal
			for line in record.materiaux_line:
				somme_materiaux += line.price_subtotal
	
		self.prix_debourse = (somme_mo + somme_materiel + somme_materiaux)
		self.price_unit = (somme_mo + somme_materiel + somme_materiaux)
		print "PRIX DEBOURSE"
		print self.prix_debourse
		print self.price_unit
		print self.rendement

	@api.depends('price_unit','product_uom_qty')
	def _compute_order_line_montant(self):
		for record in self:
			record.price_subtotal= record.price_unit * record.product_uom_qty

	@api.multi
	def import_excel(self):
		print "IMPORT EXCEL OUVRAGE ELEMENTAIRE"
		# Generating of the excel file to be read by openpyxl
		my_file = self.excel_file.decode('base64')
		excel_fileobj = TemporaryFile('wb+')
		excel_fileobj.write(my_file)
		excel_fileobj.seek(0)

		# Create workbook
		workbook = openpyxl.load_workbook(excel_fileobj, data_only=True)
		# Get the first sheet of excel file
		sheet = workbook[workbook.get_sheet_names()[0]]

		print "DISPLAYING EXCEL"
		# Iteration on each rows in excel
		for row in sheet.rows:
			for col in row:
				print col.value
		return True

# class Binary(http.Controller):
# 	@http.route('/datas', type='http', auth="public")
# 	@serialize_exception
# 	def download_document(self,model,field,id,filename=None, **kw):
# 		""" Download link for files stored as binary fields.
# 		:param str model: name of the model to fetch the binary from
# 		:param str field: binary field
# 		:param str id: id of the record from which to fetch the binary
# 		:param str filename: field holding the file's name, if any
# 		:returns: :class:`werkzeug.wrappers.Response`
# 		"""
# 		Model = request.registry[model]
# 		cr, uid, context = request.cr, request.uid, request.context
# 		fields = [field]
# 		res = Model.read(cr, uid, [int(id)], fields, context)[0]
# 		filecontent = base64.b64decode(res.get(field) or '')
# 		if not filecontent:
# 			return request.not_found()
# 		else:
# 			if not filename:
# 				filename = '%s_%s' % (model.replace('.', '_'), id)
# 				return request.make_response(filecontent,[('Content-Type', 'application/octet-stream'),('Content-Disposition', content_disposition(filename))]) 

class Coeff(models.Model):
	_name="gent.coeff"
	_rec_name = "coeff"
	frais_1 = fields.Float("Frais Généraux proportionnel au débourse", compute = "_compute_frais_1")
	frais_2 = fields.Float("Bénéfice brut et frais proportionnel au prix de revient", compute = "_compute_frais_2")
	frais_3 = fields.Float("Frais proportionnel au prix de règlement", compute = "_compute_frais_3")

	frais_11 = fields.Float("Frais d'agence et patente (%)", default=0)
	frais_12 = fields.Float("Frais de chantier (%)", default=0)
	frais_13 = fields.Float("Frais d'études et de laboratoire (%)", default=0)
	frais_14 = fields.Float("Assurances (%)", default=0)

	frais_21 = fields.Float("Bénéfice net et impôt sur le bénéfice (%)", default=20)
	frais_22 = fields.Float("Aléas techniques (%)", default=0)
	frais_23 = fields.Float("Aléas de révision de prix (%)", default=0)
	frais_24 = fields.Float("Frais financier (%)", default=0)

	tva = fields.Float("Frais financier", default=20)

	frais_31 = fields.Float("Frais pour les entreprises qui n'ont pas leur siège à Madagascar (%)", default=0)

	coeff = fields.Float("Coefficient de vente K", compute="_compute_coeff")



	@api.depends('frais_11', 'frais_12', 'frais_13', 'frais_14')
	def _compute_frais_1(self):
		for record in self:
			record.frais_1 = record.frais_11 + record.frais_12 + record.frais_13 + record.frais_14

	@api.depends('frais_21', 'frais_22', 'frais_23', 'frais_24')
	def _compute_frais_2(self):
		for record in self:
			record.frais_2 = record.frais_21 + record.frais_22 + record.frais_23 + record.frais_24

	@api.depends('frais_31')
	def _compute_frais_3(self):
		for record in self:
			record.frais_3 = record.frais_31

	@api.depends('frais_1', "frais_2", "frais_3", "tva")
	def _compute_coeff(self):
		for record in self:
			record.coeff = ( (1 + (record.frais_1/100) ) *(1 + (record.frais_2/100) ) ) / (1 - (record.frais_3*(1+(record.tva/100))))


class GentProject(models.Model):
	_inherit="project.project"
	order_id =  fields.Many2one('sale.order', 'Devis', domain=[('state', '!=', 'draft')], change_default=True, ondelete='cascade')
	tasks = fields.One2many('project.task', 'project_id', "Task Activities")
	pourcentage = fields.Float('pourcentage')
	@api.onchange('order_id')
	def on_change_order_id(self):
		print "ORDER ID CHANGE"
		print self.tasks
		# self.partner_id = self.order_id.partner_id
		# self.name = self.order_id.avantmetre.name
		self.write({
			'partner_id': self.order_id.partner_id,
			'name': self.order_id.avantmetre.name
			})
		# result = []
		# project_id = self._origin.id
		# # if(project_id):
		# for order_line in self.order_id.order_line:
		# 	result.append((0,0,{'name': order_line.name}))
		# 	self.env['project.task'].create({'name': order_line.name, 'project_id': project_id})

		# print "Writing"
		# print result
		# self.tasks = result

class GentProjectTask(models.Model):
	_inherit="project.task"
	project_id = fields.Many2one('project.project', 'Project', ondelete='set null', select=True, change_default=True)
        
	

class GentSaleLayout(models.Model):
	_inherit="sale_layout.category"



class GentStockMove(models.Model):
	_inherit="stock.move"

class GentStockPicking(models.Model):
	_inherit="stock.picking"

class GentProductProduct(models.Model):
	_inherit="product.product"
	_sql_constraints = [('name_unique', 'unique(name_template)', "Le nom de l'article devrait être unique")]

class GentProductUom(models.Model):
	_inherit="product.uom"

class GentProductPriceHistory(models.Model):
	_inherit="product.price.history"

class GentAccount(models.Model):
	_inherit = "account.invoice"

	avantmetre = fields.Float('AvantMetre')

class GentAccountInvoiceLine(models.Model):
	_inherit = "account.invoice.line"

# class GentProjectAttachement(models.Model):
# 	_name= "gent.account.attachement"
# 	_inherit="account.analytic.account"
	
# 	name = fields.Char(string="Nom")
# 	session_ids = fields.One2many('gent.attachement', 'attachement_id', string="Sessions")

class GentAttachement(models.Model):
	_name = "gent.attachement"
	# rendement = fields.Float('Rendement',default=1,required=True)
	excel_file = fields.Binary(string='Excel File')
	# date_debut = fields.date('Date de debut')
	# date_confirm = fields.date('Confirmation Date')
	projet=fields.Many2one('project.project', 'Projet', required=True, select=True)
	begin_date = fields.Date(required=True)
	end_date = fields.Date(required=True)
	excel_pourcentage = fields.Binary(string='Attachement', store=True)
	pourcentage = fields.Float('Pourcentage')
	attachement_id = fields.Many2one('account.analytic.account',ondelete='cascade', string="Attachement")

	@api.onchange('projet')
	def on_projet(self):
		print "On change project"
		
		print self.projet.order_id.name
		print "price_list"

	@api.multi
	def import_percentage(self):
		print "IMPORT EXCEL pourcentage"
		my_file = self.excel_pourcentage.decode('base64')
		excel_fileobj = TemporaryFile('wb+')
		excel_fileobj.write(my_file)
		excel_fileobj.seek(0)
		# Create workbook
		wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
		# Get the first sheet of excel file
		ws = wb[wb.get_sheet_names()[0]]
		start = False
		sections =[]
		self.rubrique_line_ids = []
		rubrique =""
		lines = []
		start_line =False
		for row in ws:
			val1 = row[0].value
			val2 = row[1].value
			val13 = row[12].value
			if val2 == "TOTAL GENERAL TTC  (en Ariary)":
				print val13
				self.projet.order_id.pourcentage = val13*100
				self.pourcentage = val13*100


class Chantier(models.Model):
	_inherit= "mrp.bom"
	bom_line_labour_ids =  fields.One2many('mrp.bom.line.labour', 'bom_id', 'BoM Lines', copy=True)
	bom_line_materiel_ids =  fields.One2many('mrp.bom.line.materiel', 'bom_id', 'BoM Lines', copy=True)

	
class gent_mrp_bom_line(models.Model):
	_inherit= "mrp.bom.line"

	price_unit = fields.Float('Unit Price', required=True, digits_compute= dp.get_precision('Product Price'), compute="_compute_pu")
	price_subtotal =  fields.Float('Montant',compute='_compute_subtotal')
	
	# budget_product_qty = fields.Float('Budget Quantité', required=True, digits_compute=dp.get_precision('Product Unit of Measure'))
	
	@api.one
	@api.depends('price_unit','product_qty')
	def _compute_subtotal(self):
		for record in self:
			record.price_subtotal = record.price_unit * record.product_qty

	@api.one
	@api.depends('product_id')
	def _compute_pu(self):
		for record in self:
			record.price_unit = record.product_id.standard_price

class gent_mrp_bom_line(models.Model):
	_inherit= "mrp.bom.line"
	_name="mrp.bom.line.labour"

class gent_mrp_bom_line(models.Model):
	_inherit= "mrp.bom.line"
	_name="mrp.bom.line.materiel"
	

